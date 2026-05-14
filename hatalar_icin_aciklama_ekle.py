#!/usr/bin/env python3

import openpyxl
import os
import sys
from glob import glob
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from datetime import datetime
import hata_kodlari

import locale
locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')

bu_yil = datetime.now().year
bu_ay = datetime.now().strftime('%B')

bu_dizin = os.path.dirname(__file__) + '/rapor/'
excel_file =glob(bu_dizin.replace('./', '') + str(bu_yil) + '/' + str(bu_ay) + '/1-maas_kontrol_raporu_v2.xlsx')

wb = load_workbook(''.join(excel_file), data_only = True)
sayfa = wb.worksheets[0]

def _get_column_letter(col_idx):
    if not 1 <= col_idx <= 50:
        raise ValueError('Hatalı sütun numarası: {0}'.format(col_idx))
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx, 26)
        if remainder == 0:
            remainder = 26
            col_idx -= 1
        letters.append(chr(remainder+64))
    return ''.join(reversed(letters))

def liste_to_yorum(metin_listesi):
    if isinstance(metin_listesi, list):
        return "\n".join(f"{i+1}. {madde}" for i, madde in enumerate(metin_listesi))
    return metin_listesi

def sutun_genislet(sayfa):
    # Tüm sütunları tek tek dönüyoruz
    for sutun_indeks in range(1, sayfa.max_column + 1):
        maks_uzunluk = 0
        
        # Sütun harfini _get_column_letter fonksiyonunla alıyoruz
        sutun_harfi = _get_column_letter(sutun_indeks)
        
        # 2. satırdan itibaren kontrol ederek en uzun veriyi buluyoruz
        for satir in range(2, sayfa.max_row + 1):
            hucre = sayfa.cell(row=satir, column=sutun_indeks)

            # Hizalama kuralını tanımlıyoruz
            ortala = Alignment(horizontal='center', vertical='center', wrap_text=False)

            hucre.alignment = ortala

            if hucre.value:
                # Verinin uzunluğunu ölçüyoruz
                mevcut_uzunluk = len(str(hucre.value))
                if mevcut_uzunluk > maks_uzunluk:
                    maks_uzunluk = mevcut_uzunluk
        
        # Sütun genişliğini ayarlıyoruz (metin tam sığsın diye bir miktar boşluk (pay) ekliyoruz)
        ayarlanmis_genislik = (maks_uzunluk + 2)
        
        # sütun genişliğini tanımlıyoruz
        sayfa.column_dimensions[sutun_harfi].width = ayarlanmis_genislik

def aciklama_ekle():
    sutun_ve_hata_kodlari = {
        'Adı Soyadı'                : liste_to_yorum(hata_kodlari.ad_soyad_hata),
        'Hizmet Sınıfı'             : liste_to_yorum(hata_kodlari.sinif_hata),
        'Görev Ünvanı'              : liste_to_yorum(hata_kodlari.unvan_hata),
        'Derece/Kademe'             : liste_to_yorum(hata_kodlari.derece_kademe_hata),
        'Gösterge/Aylık Puanı'      : liste_to_yorum(hata_kodlari.gosterge_puani_hata),
        'Gösterge/Aylık Tutarı'     : liste_to_yorum(hata_kodlari.aylik_tutar_hata),
        'Ek Gösterge Puanı'         : liste_to_yorum(hata_kodlari.ek_gosterge_hata),
        'Ek Gösterge Tutarı'        : liste_to_yorum(hata_kodlari.ek_gos_ayligi_hata),
        'Yan Ödeme Puanı'           : liste_to_yorum(hata_kodlari.yan_odeme_hata),
        'Yan Ödeme Tutarı'          : liste_to_yorum(hata_kodlari.yan_odeme_aylik_hata),
        'Ek Tazminat Puanı'         : liste_to_yorum(hata_kodlari.ek_tazminat_puani_hata),
        'Özel Hiz. Taz. Puanı'      : liste_to_yorum(hata_kodlari.ozel_hiz_taz_puani_hata),
        'Özel Hiz. Taz. Tutarı'     : liste_to_yorum(hata_kodlari.ozel_hiz_taz_tutar_hata),
        'Ek Ödeme Puanı (666 KHK)'  : liste_to_yorum(hata_kodlari.khk_666_puani_hata),
        'Ek Ödeme Tutarı (666 KHK)' : liste_to_yorum(hata_kodlari.khk_666_tutar_hata),
        'İlave Ödeme (375-40 KHK)'  : liste_to_yorum(hata_kodlari.ilave_odeme_hata),
    }

    # satır ve sütunları tek tek kontrol ediyoruz
    for satir in range(1, sayfa.max_row+1):
        for sutun in range(1, sayfa.max_column+1):
            ## Arkaplanı kırmızı olan sütunları buluyoruz
            if sayfa.cell(row=satir, column=sutun).fill.start_color.index != '00000000':
                # Başlığın hücre adresini [A1] buluyoruz
                baslik_degeri = sayfa[f'{_get_column_letter(sutun-1)}1'].value 

                # Başlık değeri, sözlüğümüzün içinde geçiyor mu?
                if baslik_degeri in sutun_ve_hata_kodlari:
                    hata_mesaji = sutun_ve_hata_kodlari[baslik_degeri]
                    
                    # Yorumu, ilgili hücreye ekliyoruz
                    sayfa.cell(row=satir, column=sutun).comment = Comment(text=f"{hata_mesaji}", author=f'yahya.yildirim@diyanet.gov.tr')

    sutun_genislet(sayfa)
    wb.save('./rapor/' + str(bu_yil) + '/' + str(bu_ay) + '/1-maas_kontrol_raporu_v2.xlsx')
    print('%100')

if __name__ == '__main__':
    aciklama_ekle()